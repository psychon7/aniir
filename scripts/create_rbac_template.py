from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ========== Sheet 1: RBAC Matrix (Main Configuration) ==========
ws_matrix = wb.active
ws_matrix.title = "RBAC Matrix"

# Define colors
header_fill = PatternFill('solid', fgColor='1F4E79')
header_font = Font(bold=True, color='FFFFFF', size=11)
bu_fills = {
    'LED': PatternFill('solid', fgColor='DBEAFE'),       # Light blue
    'HVAC': PatternFill('solid', fgColor='D1FAE5'),      # Light green
    'DOMOTICS': PatternFill('solid', fgColor='FCE7F3'),  # Light pink
    'ACCESSORIES': PatternFill('solid', fgColor='EDE9FE'),  # Light purple
    'WAVECONCEPT': PatternFill('solid', fgColor='FFEDD5'),  # Light orange
}
role_font = Font(bold=True, size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Business Units
business_units = ['LED', 'HVAC', 'DOMOTICS', 'ACCESSORIES', 'WAVECONCEPT']

# Roles with hierarchy levels
roles = [
    ('Admin', 99, 'Full system access - all permissions'),
    ('Manager', 89, 'Management level with team oversight'),
    ('Assistant', 79, 'Administrative support staff'),
    ('Comptable', 69, 'Finance and accounting'),
    ('Commercial', 59, 'Sales and commercial operations'),
]

# Modules (from existing TR_SCR_Screen)
modules = [
    ('Dashboard', 'Main dashboard and overview'),
    ('Clients', 'Client management'),
    ('Client Orders', 'Sales order processing'),
    ('Client Invoices', 'Invoice generation and management'),
    ('Quotes/Cost Plans', 'Quotation and pricing'),
    ('Products', 'Product catalog management'),
    ('Suppliers', 'Supplier management'),
    ('Supplier Orders', 'Purchase orders'),
    ('Supplier Invoices', 'Supplier invoice processing'),
    ('Projects', 'Project management'),
    ('Warehouse', 'Warehouse operations'),
    ('Inventory', 'Stock management'),
    ('Logistics', 'Delivery and logistics'),
    ('Users', 'User administration'),
    ('Settings', 'System configuration'),
    ('Reports', 'Reporting and analytics'),
]

# Access levels (permissions)
access_types = ['Read', 'Create', 'Modify', 'Delete', 'Validate', 'Cancel', 'Activate']

# Headers
ws_matrix['A1'] = 'RBAC Configuration Matrix - AX TECH ERP'
ws_matrix['A1'].font = Font(bold=True, size=16, color='1F4E79')
ws_matrix.merge_cells('A1:K1')

ws_matrix['A2'] = 'Instructions: Mark each cell with access level (✓ = Granted, ✗ = Denied, or leave blank for inherited)'
ws_matrix['A2'].font = Font(italic=True, size=10, color='666666')
ws_matrix.merge_cells('A2:K2')

# Column headers
headers = ['Business Unit', 'Role', 'Role Level', 'Module', 'Description'] + access_types
for col, header in enumerate(headers, 1):
    cell = ws_matrix.cell(row=4, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

# Add data validation for access columns
yes_no_dv = DataValidation(type="list", formula1='"✓,✗,"', allow_blank=True)
ws_matrix.add_data_validation(yes_no_dv)

# Populate matrix
row = 5
for bu in business_units:
    bu_start_row = row
    for role_name, role_level, role_desc in roles:
        for module_name, module_desc in modules:
            ws_matrix.cell(row=row, column=1, value=bu).fill = bu_fills[bu]
            ws_matrix.cell(row=row, column=1).border = thin_border
            ws_matrix.cell(row=row, column=1).alignment = center_align
            
            ws_matrix.cell(row=row, column=2, value=role_name)
            ws_matrix.cell(row=row, column=2).font = role_font
            ws_matrix.cell(row=row, column=2).border = thin_border
            ws_matrix.cell(row=row, column=2).alignment = center_align
            
            ws_matrix.cell(row=row, column=3, value=role_level)
            ws_matrix.cell(row=row, column=3).border = thin_border
            ws_matrix.cell(row=row, column=3).alignment = center_align
            
            ws_matrix.cell(row=row, column=4, value=module_name)
            ws_matrix.cell(row=row, column=4).border = thin_border
            ws_matrix.cell(row=row, column=4).alignment = left_align
            
            ws_matrix.cell(row=row, column=5, value=module_desc)
            ws_matrix.cell(row=row, column=5).border = thin_border
            ws_matrix.cell(row=row, column=5).alignment = left_align
            
            # Access level columns with validation
            for col in range(6, 6 + len(access_types)):
                cell = ws_matrix.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_align
                yes_no_dv.add(cell)
            
            row += 1

# Adjust column widths
col_widths = [15, 12, 10, 18, 35, 8, 8, 8, 8, 10, 8, 10]
for i, width in enumerate(col_widths, 1):
    ws_matrix.column_dimensions[get_column_letter(i)].width = width

# Freeze panes
ws_matrix.freeze_panes = 'F5'

# ========== Sheet 2: Business Units ==========
ws_bu = wb.create_sheet("Business Units")
ws_bu['A1'] = 'Business Unit Configuration'
ws_bu['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws_bu.merge_cells('A1:E1')

bu_headers = ['BU Code', 'BU Name', 'Color', 'Description', 'Active']
for col, header in enumerate(bu_headers, 1):
    cell = ws_bu.cell(row=3, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

bu_data = [
    ('LED', 'LED Lighting', '#3B82F6', 'LED lighting products division', 'Yes'),
    ('HVAC', 'HVAC Systems', '#10B981', 'Heating, ventilation and AC', 'Yes'),
    ('DOMOTICS', 'Domotics & Smart', '#EC4899', 'Smart home and automation', 'Yes'),
    ('ACCESSORIES', 'Accessories', '#8B5CF6', 'Product accessories', 'Yes'),
    ('WAVECONCEPT', 'WaveConcept', '#F97316', 'WaveConcept products', 'Yes'),
]

for row_idx, (code, name, color, desc, active) in enumerate(bu_data, 4):
    ws_bu.cell(row=row_idx, column=1, value=code).border = thin_border
    ws_bu.cell(row=row_idx, column=2, value=name).border = thin_border
    ws_bu.cell(row=row_idx, column=3, value=color).border = thin_border
    ws_bu.cell(row=row_idx, column=4, value=desc).border = thin_border
    ws_bu.cell(row=row_idx, column=5, value=active).border = thin_border

for i, w in enumerate([12, 20, 12, 35, 10], 1):
    ws_bu.column_dimensions[get_column_letter(i)].width = w

# ========== Sheet 3: Roles ==========
ws_roles = wb.create_sheet("Roles")
ws_roles['A1'] = 'Role Configuration'
ws_roles['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws_roles.merge_cells('A1:G1')

role_headers = ['Role ID', 'Role Name', 'Level', 'Parent Role', 'System Role', 'Description', 'Notes']
for col, header in enumerate(role_headers, 1):
    cell = ws_roles.cell(row=3, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

role_data = [
    (1, 'Admin', 99, '', 'Yes', 'Full system administrator', 'Cannot be modified'),
    (5, 'Manager', 89, 'Admin', 'Yes', 'Management level with team oversight', 'Inherits from Admin'),
    (2, 'Assistant', 79, 'Manager', 'Yes', 'Administrative support staff', ''),
    (4, 'Comptable', 69, 'Manager', 'Yes', 'Finance and accounting', ''),
    (3, 'Commercial', 59, 'Manager', 'Yes', 'Sales and commercial operations', ''),
]

for row_idx, data in enumerate(role_data, 4):
    for col_idx, val in enumerate(data, 1):
        cell = ws_roles.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = center_align if col_idx < 6 else left_align

for i, w in enumerate([10, 15, 8, 15, 12, 35, 25], 1):
    ws_roles.column_dimensions[get_column_letter(i)].width = w

# ========== Sheet 4: Modules ==========
ws_modules = wb.create_sheet("Modules")
ws_modules['A1'] = 'Module/Screen Configuration'
ws_modules['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws_modules.merge_cells('A1:F1')

mod_headers = ['Screen ID', 'Module Name', 'API Resource', 'API Endpoint', 'Parent Module', 'Description']
for col, header in enumerate(mod_headers, 1):
    cell = ws_modules.cell(row=3, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

mod_data = [
    (7, 'Clients', 'clients', '/api/v1/clients', '', 'Client management'),
    (12, 'Client Orders', 'orders', '/api/v1/orders', '', 'Sales order processing'),
    (10, 'Client Invoices', 'invoices', '/api/v1/invoices', '', 'Invoice management'),
    (17, 'Quotes/Cost Plans', 'quotes', '/api/v1/cost-plans', '', 'Quotation and pricing'),
    (26, 'Products', 'products', '/api/v1/products', '', 'Product catalog'),
    (40, 'Suppliers', 'suppliers', '/api/v1/suppliers', '', 'Supplier management'),
    (46, 'Supplier Orders', 'supplier-orders', '/api/v1/supplier-orders', '', 'Purchase orders'),
    (44, 'Supplier Invoices', 'supplier-invoices', '/api/v1/supplier-invoices', '', 'Supplier invoices'),
    (32, 'Projects', 'projects', '/api/v1/projects', '', 'Project management'),
    (51, 'Warehouse', 'warehouse', '/api/v1/warehouse', '', 'Warehouse operations'),
    (48, 'Inventory', 'inventory', '/api/v1/inventory', '', 'Stock management'),
    (23, 'Logistics', 'logistics', '/api/v1/logistics', '', 'Delivery and logistics'),
    (47, 'Users', 'users', '/api/v1/users', '', 'User administration'),
    (2, 'Settings', 'admin', '/api/v1/admin/settings', '', 'System configuration'),
]

for row_idx, data in enumerate(mod_data, 4):
    for col_idx, val in enumerate(data, 1):
        cell = ws_modules.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = center_align if col_idx < 4 else left_align

for i, w in enumerate([10, 18, 18, 25, 15, 30], 1):
    ws_modules.column_dimensions[get_column_letter(i)].width = w

# ========== Sheet 5: Permission Legend ==========
ws_legend = wb.create_sheet("Permission Legend")
ws_legend['A1'] = 'Permission Types & API Mapping'
ws_legend['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws_legend.merge_cells('A1:D1')

legend_headers = ['Permission', 'DB Column', 'HTTP Methods', 'Description']
for col, header in enumerate(legend_headers, 1):
    cell = ws_legend.cell(row=3, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

legend_data = [
    ('Read', 'rit_read', 'GET', 'View screen/data'),
    ('Create', 'rit_create', 'POST', 'Create new records'),
    ('Modify', 'rit_modify', 'PUT, PATCH', 'Edit existing records'),
    ('Delete', 'rit_delete', 'DELETE', 'Delete records'),
    ('Validate', 'rit_valid', 'POST /approve', 'Approve/Validate records'),
    ('Cancel', 'rit_cancel', 'POST /cancel', 'Cancel/Void records'),
    ('Activate', 'rit_active', 'PATCH /activate', 'Activate/Deactivate'),
    ('Super Right', 'rit_super_right', 'ALL', 'Full access - bypass all checks'),
]

for row_idx, data in enumerate(legend_data, 4):
    for col_idx, val in enumerate(data, 1):
        cell = ws_legend.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = left_align

for i, w in enumerate([15, 18, 18, 35], 1):
    ws_legend.column_dimensions[get_column_letter(i)].width = w

# ========== Sheet 6: Sample Users ==========
ws_users = wb.create_sheet("Sample Users")
ws_users['A1'] = 'User-Role-Business Unit Assignment Examples'
ws_users['A1'].font = Font(bold=True, size=14, color='1F4E79')
ws_users.merge_cells('A1:F1')

user_headers = ['User Name', 'Email', 'Role', 'Business Units', 'Primary BU', 'Notes']
for col, header in enumerate(user_headers, 1):
    cell = ws_users.cell(row=3, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

user_data = [
    ('Alice Admin', 'alice@axtech.com', 'Admin', 'ALL', 'LED', 'Full system access'),
    ('Bob Manager', 'bob@axtech.com', 'Manager', 'LED, HVAC', 'LED', 'Manages LED and HVAC'),
    ('Carol Sales', 'carol@axtech.com', 'Commercial', 'LED', 'LED', 'LED sales only'),
    ('David Finance', 'david@axtech.com', 'Comptable', 'ALL', 'LED', 'All BU finance access'),
    ('Eve Assistant', 'eve@axtech.com', 'Assistant', 'DOMOTICS', 'DOMOTICS', 'Domotics support'),
]

for row_idx, data in enumerate(user_data, 4):
    for col_idx, val in enumerate(data, 1):
        cell = ws_users.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = left_align

for i, w in enumerate([18, 25, 12, 20, 12, 30], 1):
    ws_users.column_dimensions[get_column_letter(i)].width = w

# ========== Sheet 7: Notes & Discussion ==========
ws_notes = wb.create_sheet("Notes & Discussion")
ws_notes['A1'] = 'RBAC Discussion Notes'
ws_notes['A1'].font = Font(bold=True, size=14, color='1F4E79')

notes = [
    '',
    'Key Questions to Discuss:',
    '',
    '1. Business Unit Isolation:',
    '   - Should users from LED see HVAC data?',
    '   - Can a user belong to multiple business units?',
    '   - Should cross-BU reporting be allowed for managers?',
    '',
    '2. Role Hierarchy:',
    '   - Should Manager inherit all Admin permissions or specific subset?',
    '   - Can custom roles be created per business unit?',
    '   - Is role level (99, 89, etc.) sufficient for hierarchy?',
    '',
    '3. Module Access:',
    '   - Which modules are BU-specific vs global?',
    '   - Should Finance access be cross-BU by default?',
    '   - Are there modules that should be Admin-only?',
    '',
    '4. Permission Granularity:',
    '   - Is Read/Create/Modify/Delete/Validate/Cancel/Activate sufficient?',
    '   - Do we need field-level permissions (e.g., hide sensitive prices)?',
    '   - Should some permissions be tied together (e.g., Modify implies Read)?',
    '',
    '5. Default Permissions:',
    '   - What should new users get by default?',
    '   - Should there be a "Guest" or "Read-Only" role?',
    '',
    '6. Audit Requirements:',
    '   - What actions need to be logged?',
    '   - How long should audit logs be retained?',
    '',
    'Decision Log:',
    '------------',
    '(Add decisions here as they are made)',
]

for row_idx, note in enumerate(notes, 3):
    ws_notes.cell(row=row_idx, column=1, value=note)
    if note.endswith(':') and not note.startswith(' '):
        ws_notes.cell(row=row_idx, column=1).font = Font(bold=True)

ws_notes.column_dimensions['A'].width = 80

# Save workbook
output_path = '/Users/mohankumarv/Desktop/Projects/Clients/AXTECH/ERP2025/RBAC_Configuration_Template.xlsx'
wb.save(output_path)
print(f"Created: {output_path}")
